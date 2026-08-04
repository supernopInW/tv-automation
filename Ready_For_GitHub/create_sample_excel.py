"""
สร้างไฟล์ Excel ตัวอย่างแผนปฏิบัติงาน T&V
ครอบคลุมทุกประเด็นงานและ edge case สำหรับทดสอบระบบ
"""
import sys
try:
    sys.stdout.reconfigure(errors='replace')
    sys.stderr.reconfigure(errors='replace')
except AttributeError:
    pass

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os

OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ตัวอย่างแผนงาน_มิย69.xlsx")

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════
# Style Definitions
# ═══════════════════════════════════════════════════════════
FONT_TITLE = Font(name="TH SarabunPSK", size=16, bold=True)
FONT_HEADER = Font(name="TH SarabunPSK", size=13, bold=True, color="FFFFFF")
FONT_SUBHEADER = Font(name="TH SarabunPSK", size=12, bold=True)
FONT_BODY = Font(name="TH SarabunPSK", size=12)
FONT_NOTE = Font(name="TH SarabunPSK", size=11, italic=True, color="888888")
FONT_OVERRIDE = Font(name="TH SarabunPSK", size=11, color="2563EB")

FILL_HEADER = PatternFill(start_color="2D5F2D", end_color="2D5F2D", fill_type="solid")
FILL_SUBHEADER = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
FILL_HOLIDAY = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
FILL_WEEKEND = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
FILL_OVERRIDE = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ═══════════════════════════════════════════════════════════
# Sheet: มิ.ย.69
# ═══════════════════════════════════════════════════════════
ws = wb.active
ws.title = "มิ.ย.69"

# Column widths (A=ลำดับ, B=วันที่, C=กิจกรรม, D=เครื่องมือ, E=สถานที่, F=เป้าหมาย, G=ผู้ร่วม, H=ตำบล, I-M=Override)
col_widths = [6, 12, 45, 25, 25, 18, 22, 20, 12, 12, 18, 20, 12]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Row 1: Title ──
ws.merge_cells("A1:M1")
cell = ws["A1"]
cell.value = "แผนปฏิบัติงานรายเดือน ประจำเดือน มิถุนายน 2569  |  สำนักงานเกษตรอำเภอสีดา  |  ผู้รับผิดชอบ: ตำบลหนองตาดใหญ่"
cell.font = FONT_TITLE
cell.alignment = Alignment(horizontal="center", vertical="center")

# ── Row 2: Sub-header ──
ws.merge_cells("A2:H2")
cell = ws["A2"]
cell.value = "ข้อมูลหลัก (ระบบอ่านอัตโนมัติ)"
cell.font = FONT_SUBHEADER
cell.fill = FILL_SUBHEADER
cell.alignment = ALIGN_CENTER

ws.merge_cells("I2:M2")
cell = ws["I2"]
cell.value = "Override (ถ้าต้องการบังคับค่าเอง — ระบบ rules-based เท่านั้น)"
cell.font = Font(name="TH SarabunPSK", size=11, bold=True, color="2563EB")
cell.fill = FILL_OVERRIDE
cell.alignment = ALIGN_CENTER

# ── Row 3: Column explanation ──
ws.merge_cells("A3:M3")
cell = ws["A3"]
cell.value = "หมายเหตุ: ระบบจะอ่านข้อมูลตั้งแต่แถวที่ 6 เป็นต้นไป (แถว 1-5 เป็น Header) | คอลัมน์ A=ลำดับ, B=วันที่, C=กิจกรรม, D=เครื่องมือ/วิธีการ, E=สถานที่, F=เป้าหมาย, G=ผู้ร่วมงาน, H=ตำบล"
cell.font = FONT_NOTE
cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Row 4: Sub-note ──
ws.merge_cells("A4:M4")
cell = ws["A4"]
cell.value = "คอลัมน์ I-M (Override): I=issue_val, J=activity_val, K=other_text, L=location, M=target_num — ใช้เฉพาะ parser แบบ rules-based (ไม่มี Gemini API Key)"
cell.font = FONT_NOTE
cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Row 5: Header row ──
headers = [
    "ลำดับ",         # A - iloc[0]
    "วันที่",        # B - iloc[1]
    "กิจกรรม / เนื้องาน",  # C - iloc[2]
    "เครื่องมือ / วิธีการ", # D - iloc[3]
    "สถานที่",       # E - iloc[4]
    "เป้าหมาย",      # F - iloc[5]
    "ผู้ร่วมดำเนินงาน",    # G - iloc[6]
    "ตำบล",          # H - iloc[7]
    "issue_val",     # I - iloc[8]  (override)
    "activity_val",  # J - iloc[9]  (override)
    "other_text",    # K - iloc[10] (override)
    "location",      # L - iloc[11] (override)
    "target_num",    # M - iloc[12] (override)
]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=5, column=col_idx, value=header)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER

# ═══════════════════════════════════════════════════════════
# Data Rows — ครอบคลุมทุกกรณี
# ═══════════════════════════════════════════════════════════
# Format: [ลำดับ, วันที่, กิจกรรม, เครื่องมือ, สถานที่, เป้าหมาย, ผู้ร่วม, ตำบล, *override(I-M)]
# override columns are optional — leave as None to skip

data_rows = [
    # ─────────────────────────────────────────────────────────
    # ประเด็น 1: การถ่ายทอดความรู้ (TRAINING/MEETING)
    # ─────────────────────────────────────────────────────────
    [1, "2 มิ.ย.", "ประชุมสำนักงานเกษตรอำเภอประจำสัปดาห์ (WM)",
     "ประชุม", "สนง.กษอ.สีดา", "จนท.", "นายสมศักดิ์ ใจดี", "หนองตาดใหญ่",
     None, None, None, None, None],

    [2, "5 มิ.ย.", "ประชุมประจำเดือน (MM) เกษตรอำเภอสีดา",
     "ประชุม", "สนง.กษอ.สีดา", "จนท. ทุกท่าน", "นายสมศักดิ์ ใจดี", "หนองตาดใหญ่",
     None, None, None, None, None],

    [3, "9 มิ.ย.", "ประชุมสำนักงานเกษตรอำเภอ (DM) ชี้แจงแนวทางปฏิบัติงานประจำเดือน",
     "ประชุม/ชี้แจง", "สนง.กษอ.สีดา", "เจ้าหน้าที่", "นายสมศักดิ์ ใจดี", "หนองตาดใหญ่",
     None, None, None, None, None],

    [4, "12 มิ.ย.", "อบรมเกษตรกรระดับอำเภอ เรื่องการป้องกันศัตรูพืช (DW)",
     "อบรม/สาธิต", "สนง.กษอ.สีดา", "เกษตรกร 30 ราย", "นางสาวพิมพ์ใจ รักเรียน", "หนองตาดใหญ่",
     None, None, None, None, None],

    [5, "16 มิ.ย.", "ประชุมชี้แจงโครงการส่งเสริมปลูกข้าวโพด",
     "ประชุม/ชี้แจง", "สนง.กษอ.สีดา", "เกษตรกร 20 คน", "นายสมศักดิ์ ใจดี", "หนองตาดใหญ่",
     None, None, None, None, None],

    # ─────────────────────────────────────────────────────────
    # ประเด็น 2: การเยี่ยมเยียน (VISITING)
    # ─────────────────────────────────────────────────────────
    [6, "3 มิ.ย.", "เยี่ยมเยียนศูนย์เรียนรู้การเพิ่มประสิทธิภาพการผลิตสินค้าเกษตร (ศพก.)",
     "เยี่ยมเยียน/สำรวจ", "ศพก.ตำบลสีดา", "เกษตรกร 15 ราย", "นายสมศักดิ์ ใจดี", "สีดา",
     None, None, None, None, None],

    [7, "4 มิ.ย.", "ติดตามเกษตรแปลงใหญ่ข้าว ตรวจสอบการเตรียมแปลง",
     "เยี่ยมเยียน", "ตำบลหนองตาดใหญ่", "สมาชิกแปลงใหญ่ 25 ราย", "นางสาวพิมพ์ใจ รักเรียน", "หนองตาดใหญ่",
     None, None, None, None, None],

    [8, "6 มิ.ย.", "ประชุมกลุ่มวิสาหกิจชุมชนแปรรูปผลผลิตทางการเกษตร",
     "ประชุม/แนะนำ", "ตำบลสามเมือง", "สมาชิก วสช. 12 คน", "นายสมศักดิ์ ใจดี", "สามเมือง",
     None, None, None, None, None],

    [9, "10 มิ.ย.", "เยี่ยมเยียน Smart Farmer ต้นแบบ ด้านการปลูกข้าวหอมมะลิ",
     "เยี่ยมเยียน/สัมภาษณ์", "บ้านโคกเขา ม.5 ตำบลหนองตาดใหญ่", "เกษตรกร 3 ราย", "นางสาวพิมพ์ใจ รักเรียน", "หนองตาดใหญ่",
     None, None, None, None, None],

    [10, "11 มิ.ย.", "ส่งเสริมเกษตรอินทรีย์และ GAP กลุ่มผลิตข้าวปลอดสาร",
     "เยี่ยมเยียน/แนะนำ", "ตำบลโนนประดู่", "เกษตรกร 10 ราย", "นายสมศักดิ์ ใจดี", "โนนประดู่",
     None, None, None, None, None],

    [11, "13 มิ.ย.", "ลงพื้นที่ส่งเสริมการปลูกพืชหลังนาตามนโยบาย Zoning by Agri-Map",
     "เยี่ยมเยียน/สำรวจ", "ตำบลหนองตาดใหญ่", "เกษตรกร 8 ราย", "นางสาวพิมพ์ใจ รักเรียน", "หนองตาดใหญ่",
     None, None, None, None, None],

    [12, "17 มิ.ย.", "เยี่ยมเยียนกลุ่มเกษตรกรผู้ปลูกข้าว ตำบลโพนทอง",
     "เยี่ยมเยียน", "ตำบลโพนทอง", "สมาชิกกลุ่ม 18 คน", "นายสมศักดิ์ ใจดี", "โพนทอง",
     None, None, None, None, None],

    [13, "18 มิ.ย.", "ติดตามโครงการเศรษฐกิจพอเพียงอันเนื่องมาจากพระราชดำริ",
     "เยี่ยมเยียน/ติดตาม", "ตำบลสามเมือง", "เกษตรกร 5 ราย", "นางสาวพิมพ์ใจ รักเรียน", "สามเมือง",
     None, None, None, None, None],

    [14, "19 มิ.ย.", "จัดงานวันถ่ายทอดเทคโนโลยี (Field Day) ด้านการจัดการศัตรูพืช",
     "สาธิต/ฝึกอบรม", "ศพก.ตำบลสีดา", "เกษตรกร 50 ราย", "จนท.ทุกคน", "สีดา",
     None, None, None, None, None],

    [15, "20 มิ.ย.", "ลงพื้นที่รณรงค์ไม่เผาตอซังในพื้นที่เกษตร",
     "ชี้แจง/ประชาสัมพันธ์", "ตำบลหนองตาดใหญ่", "เกษตรกร 20 ราย", "นายสมศักดิ์ ใจดี", "หนองตาดใหญ่",
     None, None, None, None, None],

    [16, "23 มิ.ย.", "ประชุมกลุ่มแม่บ้านเกษตรกร เรื่องการแปรรูปผลผลิต",
     "ประชุม/สาธิต", "ตำบลหนองตาดใหญ่", "แม่บ้านเกษตรกร 15 คน", "นางสาวพิมพ์ใจ รักเรียน", "หนองตาดใหญ่",
     None, None, None, None, None],

    [17, "24 มิ.ย.", "ติดตามสถานการณ์น้ำท่วมพื้นที่การเกษตร หลังฝนตกหนัก",
     "สำรวจ/ติดตาม", "ตำบลโนนประดู่", "เกษตรกร 10 ราย", "นายสมศักดิ์ ใจดี", "โนนประดู่",
     None, None, None, None, None],

    # ─────────────────────────────────────────────────────────
    # ประเด็น 3: การสนับสนุน (SUPPORTING)
    # ─────────────────────────────────────────────────────────
    [18, "25 มิ.ย.", "สนับสนุนวัสดุอุปกรณ์การเกษตร แก่เกษตรกรผู้ประสบภัย",
     "สนับสนุน/แจกจ่าย", "ตำบลหนองตาดใหญ่", "เกษตรกร 30 ราย", "จนท. 3 คน", "หนองตาดใหญ่",
     None, None, None, None, None],

    # ─────────────────────────────────────────────────────────
    # ประเด็น 4: การนิเทศงาน (SUPERVISION)
    # ─────────────────────────────────────────────────────────
    [19, "26 มิ.ย.", "รับนิเทศงานจากเกษตรจังหวัดนครราชสีมา ตรวจเยี่ยมราชการ",
     "นิเทศ/ตรวจเยี่ยม", "สนง.กษอ.สีดา", "จนท.", "ทีมนิเทศจังหวัด", "หนองตาดใหญ่",
     None, None, None, None, None],

    # ─────────────────────────────────────────────────────────
    # ประเด็น 5: การจัดการข้อมูล (DATA MANAGEMENT)
    # ─────────────────────────────────────────────────────────
    [20, "4 มิ.ย.", "ปรับปรุงทะเบียนเกษตรกร (ทบก.) ประจำเดือน มิถุนายน",
     "บันทึกข้อมูล", "สนง.กษอ.สีดา", "", "นางสาวพิมพ์ใจ รักเรียน", "หนองตาดใหญ่",
     None, None, None, None, None],

    [21, "27 มิ.ย.", "จัดทำแผนพัฒนาการเกษตร ระดับตำบล ปี 2569",
     "จัดทำข้อมูล", "สนง.กษอ.สีดา", "", "นายสมศักดิ์ ใจดี", "หนองตาดใหญ่",
     None, None, None, None, None],

    # ─────────────────────────────────────────────────────────
    # Edge Cases — วันหยุด / ลา / ว่าง
    # ─────────────────────────────────────────────────────────
    [22, "1 มิ.ย.", "วันวิสาขบูชา (หยุดราชการ)",
     "", "", "", "", "หนองตาดใหญ่",
     None, None, None, None, None],

    [23, "7 มิ.ย.", "หยุดราชการ (เสาร์)",
     "", "", "", "", "หนองตาดใหญ่",
     None, None, None, None, None],

    [24, "8 มิ.ย.", "หยุดราชการ (อาทิตย์)",
     "", "", "", "", "หนองตาดใหญ่",
     None, None, None, None, None],

    [25, "14 มิ.ย.", "ลาพักผ่อน",
     "", "", "", "", "หนองตาดใหญ่",
     None, None, None, None, None],

    [26, "15 มิ.ย.", "ลาป่วย",
     "", "", "", "", "หนองตาดใหญ่",
     None, None, None, None, None],

    # ─────────────────────────────────────────────────────────
    # Edge Cases — location ผิดรูปแบบ
    # ─────────────────────────────────────────────────────────
    [27, "21 มิ.ย.", "ลงพื้นที่เยี่ยมเยียนเกษตรกรรายย่อย",
     "แผ่นพับ/ชี้แจง", "ชี้แจง/ประชาสัมพันธ์", "เกษตรกร 5 ราย", "นายสมศักดิ์ ใจดี", "หนองตาดใหญ่",
     None, None, None, None, None],

    [28, "22 มิ.ย.", "สำรวจพื้นที่การเกษตรเพื่อวางแผนปลูก",
     "สำรวจ", "ตำบล….....", "เกษตรกร 3-5 ราย", "นางสาวพิมพ์ใจ รักเรียน", "สามเมือง",
     None, None, None, None, None],

    # ─────────────────────────────────────────────────────────
    # Edge Cases — target ซับซ้อน
    # ─────────────────────────────────────────────────────────
    [29, "28 มิ.ย.", "ลงพื้นที่ติดตามสถานการณ์ภัยแล้ง",
     "เยี่ยมเยียน", "ตำบลโพนทอง", "จนท 4 คน + เกษตรกร 10 ราย", "จนท.ทุกคน", "โพนทอง",
     None, None, None, None, None],

    [30, "29 มิ.ย.", "ประชุมกลุ่มเกษตรกรทำนาแปลงใหญ่ข้าว ครั้งที่ 3/2569",
     "ประชุม/แนะนำ", "ตำบลหนองตาดใหญ่", "ทุกคน", "นายสมศักดิ์ ใจดี", "หนองตาดใหญ่",
     None, None, None, None, None],

    [31, "30 มิ.ย.", "จัดทำรายงานสรุปผลการปฏิบัติงานประจำเดือน",
     "จัดทำเอกสาร", "สนง.กษอ.สีดา", "", "นางสาวพิมพ์ใจ รักเรียน", "หนองตาดใหญ่",
     None, None, None, None, None],
]

# Write data rows (starting at Excel row 6 = pandas index 4 when header is row 5)
for row_idx, row_data in enumerate(data_rows, 6):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = FONT_BODY
        cell.alignment = ALIGN_LEFT if col_idx >= 3 else ALIGN_CENTER
        cell.border = THIN_BORDER

        # Highlight holidays/leave
        activity = row_data[2] if len(row_data) > 2 else ""
        if activity and any(kw in str(activity) for kw in ["หยุด", "วันวิสาขบูชา", "วันจักรี"]):
            cell.fill = FILL_HOLIDAY
        elif activity and any(kw in str(activity) for kw in ["ลาพักผ่อน", "ลาป่วย", "ลากิจ"]):
            cell.fill = FILL_WEEKEND

        # Highlight override columns (I-M)
        if col_idx >= 9 and value is not None:
            cell.font = FONT_OVERRIDE
            cell.fill = FILL_OVERRIDE

# ═══════════════════════════════════════════════════════════
# Sheet 2: คำอธิบายรหัส (Reference)
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("คำอธิบายรหัส")
ws2.column_dimensions["A"].width = 15
ws2.column_dimensions["B"].width = 15
ws2.column_dimensions["C"].width = 45
ws2.column_dimensions["D"].width = 50

ref_headers = ["issue_val", "activity_val", "ชื่อกิจกรรม", "คำสำคัญ / Keywords"]
for col_idx, h in enumerate(ref_headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER

ref_data = [
    # Issue 1
    ["1", "14", "ประชุมสัปดาห์ (Weekly Meeting: WM)", "WM, ประชุมสัปดาห์, ประชุมประจำสัปดาห์"],
    ["1", "13", "ประชุมอำเภอ (District Meeting: DM)", "DM, ประชุมอำเภอ, ประชุมสำนักงานเกษตรอำเภอ"],
    ["1", "12", "ประชุมเดือน (Monthly Meeting: MM)", "MM, ประชุมเดือน, ประชุมประจำเดือน"],
    ["1", "11", "ประชุมจังหวัด (Provincial Meeting: PM)", "PM, ประชุมจังหวัด"],
    ["1", "1", "สัมมนาระดับชาติ (NW)", "NW, สัมมนาระดับชาติ, ประชุมกรม"],
    ["1", "6", "สัมมนาระดับเขต (RW)", "RW, สัมมนาเขต"],
    ["1", "7", "สัมมนาระดับจังหวัด (PW)", "PW, อบรมจังหวัด"],
    ["1", "8", "สัมมนาระดับอำเภอ (DW)", "DW, อบรมอำเภอ"],
    ["1", "999", "อื่นๆ (ประชุมทั่วไป)", "ประชุมชี้แจง, ประชุมนอกรอบ"],
    # Issue 2
    ["2", "2", "ศพก. (ศูนย์เรียนรู้ฯ)", "ศพก, ศูนย์เรียนรู้"],
    ["2", "15", "เกษตรแปลงใหญ่", "แปลงใหญ่, นาแปลงใหญ่"],
    ["2", "16", "Smart Farmer / YSF", "Smart Farmer, SF, YSF, สมาร์ทฟาร์มเมอร์"],
    ["2", "17", "Zoning by Agri-Map", "Zoning, Agri-Map"],
    ["2", "18", "โครงการพระราชดำริ", "พระราชดำริ, เศรษฐกิจพอเพียง"],
    ["2", "19", "วิสาหกิจชุมชน", "วิสาหกิจชุมชน, วสช"],
    ["2", "20", "กลุ่มเกษตรกร / แม่บ้าน", "กลุ่มเกษตรกร, กลุ่มแม่บ้าน, องค์กรเกษตรกร, 3ก"],
    ["2", "22", "เกษตรอินทรีย์", "อินทรีย์, GAP, 5 ดี, เกษตรปลอดภัย"],
    ["2", "24", "พัฒนาคุณภาพสินค้าเกษตร / สุขภาพพืช", "คุณภาพสินค้า, สุขภาพพืช, ศัตรูพืช, IPM"],
    ["2", "25", "บริหารจัดการทรัพยากรน้ำ", "ทรัพยากรน้ำ, ชลประทาน"],
    ["2", "999", "อื่นๆ (เยี่ยมเยียนทั่วไป)", "ลงพื้นที่, สำรวจ, ติดตามงาน, รณรงค์, Field Day"],
    # Issue 3
    ["3", "3", "ด้านโครงสร้างและอุปกรณ์", "อุปกรณ์, สนับสนุนวัสดุ"],
    ["3", "33", "เพิ่มสมรรถนะ/ขวัญกำลังใจ", "สร้างขวัญ, กำลังใจ"],
    ["3", "34", "ด้านวิชาการ", "สนับสนุนวิชาการ"],
    # Issue 4
    ["4", "999", "นิเทศงาน (อื่นๆ)", "นิเทศ, ตรวจเยี่ยมราชการ, ตรวจติดตาม"],
    # Issue 5
    ["5", "4", "ด้านข้อมูลสารสนเทศ", "ทะเบียนเกษตรกร, ทบก, ขึ้นทะเบียน, บันทึกข้อมูล"],
    ["5", "36", "ด้านแผนพัฒนาการเกษตร", "แผนพัฒนา, จัดทำแผน"],
    ["5", "999", "อื่นๆ (จัดการข้อมูล)", "งานสารบรรณ, งานธุรการ, จัดทำรายงาน"],
]

for row_idx, row_data in enumerate(ref_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.font = FONT_BODY
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER

# ═══════════════════════════════════════════════════════════
# Sheet 3: ผลลัพธ์ที่คาดหวัง (Expected Output)
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("ผลลัพธ์ที่คาดหวัง")
exp_cols = [
    ("A", 8), ("B", 12), ("C", 40), ("D", 14), ("E", 14), ("F", 18),
    ("G", 14), ("H", 25), ("I", 14),
]
for col_letter, width in exp_cols:
    ws3.column_dimensions[col_letter].width = width

exp_headers = [
    "ลำดับ", "วันที่", "กิจกรรม (สรุป)",
    "should_include", "issue_val", "activity_val",
    "other_text", "location (ที่ระบบควรกรอก)", "target_num"
]
for col_idx, h in enumerate(exp_headers, 1):
    cell = ws3.cell(row=1, column=col_idx, value=h)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER

expected = [
    [1,  "2 มิ.ย.",  "ประชุม WM",                     True,  "1", "14",  "",                  "สนง.กษอ.สีดา",        7],
    [2,  "5 มิ.ย.",  "ประชุม MM",                      True,  "1", "12",  "",                  "สนง.กษอ.สีดา",        7],
    [3,  "9 มิ.ย.",  "ประชุม DM",                      True,  "1", "13",  "",                  "สนง.กษอ.สีดา",        7],
    [4,  "12 มิ.ย.", "อบรม DW",                        True,  "1", "8",   "",                  "สนง.กษอ.สีดา",        30],
    [5,  "16 มิ.ย.", "ประชุมชี้แจง (ทั่วไป)",           True,  "1", "999", "ชี้แจงโครงการข้าวโพด", "สนง.กษอ.สีดา",     20],
    [6,  "3 มิ.ย.",  "เยี่ยม ศพก.",                    True,  "2", "2",   "",                  "ศพก.ตำบลสีดา",        15],
    [7,  "4 มิ.ย.",  "ติดตาม แปลงใหญ่",               True,  "2", "15",  "",                  "ตำบลหนองตาดใหญ่",     25],
    [8,  "6 มิ.ย.",  "ประชุม วิสาหกิจชุมชน",           True,  "2", "19",  "",                  "ตำบลสามเมือง",         12],
    [9,  "10 มิ.ย.", "เยี่ยม Smart Farmer",            True,  "2", "16",  "",                  "บ้านโคกเขา ม.5 ตำบลหนองตาดใหญ่", 3],
    [10, "11 มิ.ย.", "เกษตรอินทรีย์ / GAP",            True,  "2", "22",  "",                  "ตำบลโนนประดู่",        10],
    [11, "13 มิ.ย.", "Zoning by Agri-Map",             True,  "2", "17",  "",                  "ตำบลหนองตาดใหญ่",     8],
    [12, "17 มิ.ย.", "เยี่ยม กลุ่มเกษตรกร",            True,  "2", "20",  "",                  "ตำบลโพนทอง",           18],
    [13, "18 มิ.ย.", "โครงการพระราชดำริ",               True,  "2", "18",  "",                  "ตำบลสามเมือง",         5],
    [14, "19 มิ.ย.", "Field Day / ศัตรูพืช",           True,  "2", "999", "งาน Field Day",     "ศพก.ตำบลสีดา",         50],
    [15, "20 มิ.ย.", "รณรงค์ไม่เผาตอซัง",              True,  "2", "999", "รณรงค์สิ่งแวดล้อม", "ตำบลหนองตาดใหญ่",     20],
    [16, "23 มิ.ย.", "ประชุม กลุ่มแม่บ้าน",             True,  "2", "20",  "",                  "ตำบลหนองตาดใหญ่",     15],
    [17, "24 มิ.ย.", "ติดตามน้ำท่วม",                  True,  "2", "999", "ติดตามน้ำท่วม",     "ตำบลโนนประดู่",        10],
    [18, "25 มิ.ย.", "สนับสนุนวัสดุ",                  True,  "3", "3",   "",                  "ตำบลหนองตาดใหญ่",     30],
    [19, "26 มิ.ย.", "นิเทศงาน",                      True,  "4", "999", "",                  "สนง.กษอ.สีดา",         7],
    [20, "4 มิ.ย.",  "ทบก. / ทะเบียนเกษตรกร",         True,  "5", "4",   "",                  "สนง.กษอ.สีดา",         0],
    [21, "27 มิ.ย.", "จัดทำแผนพัฒนาเกษตร",             True,  "5", "36",  "",                  "สนง.กษอ.สีดา",         0],
    [22, "1 มิ.ย.",  "วันวิสาขบูชา (หยุด)",             False, "2", "999", "",                  "",                     0],
    [23, "7 มิ.ย.",  "หยุด เสาร์",                     False, "2", "999", "",                  "",                     0],
    [24, "8 มิ.ย.",  "หยุด อาทิตย์",                   False, "2", "999", "",                  "",                     0],
    [25, "14 มิ.ย.", "ลาพักผ่อน",                      False, "2", "999", "",                  "",                     0],
    [26, "15 มิ.ย.", "ลาป่วย",                        False, "2", "999", "",                  "",                     0],
    [27, "21 มิ.ย.", "เยี่ยมเยียน (location ผิด→ใช้ tambon)", True, "2", "999", "เยี่ยมเยียนรายย่อย", "ตำบลหนองตาดใหญ่", 5],
    [28, "22 มิ.ย.", "สำรวจ (location มีจุดไข่ปลา→ใช้ tambon)", True, "2", "999", "สำรวจพื้นที่เกษตร", "ตำบลสามเมือง", 5],
    [29, "28 มิ.ย.", "ติดตามภัยแล้ง (target รวม)", True, "2", "999", "ติดตามภัยแล้ง",     "ตำบลโพนทอง",           14],
    [30, "29 มิ.ย.", "ประชุม+แปลงใหญ่ → issue 2",      True,  "2", "15",  "",                  "ตำบลหนองตาดใหญ่",     7],
    [31, "30 มิ.ย.", "จัดทำรายงาน",                    True,  "5", "999", "สรุปผลปฏิบัติงาน",  "สนง.กษอ.สีดา",         0],
]

for row_idx, row_data in enumerate(expected, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.font = FONT_BODY
        cell.alignment = ALIGN_LEFT if col_idx in [3, 7, 8] else ALIGN_CENTER
        cell.border = THIN_BORDER
        # Color false rows
        if col_idx == 4 and value is False:
            cell.fill = FILL_HOLIDAY

# ═══════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════
wb.save(OUTPUT_PATH)
print(f"[OK] Created: {OUTPUT_PATH}")
print(f"     Size: {os.path.getsize(OUTPUT_PATH):,} bytes")
print(f"     Sheet 1: data (31 rows, 5 issue types + edge cases)")
print(f"     Sheet 2: code reference table")
print(f"     Sheet 3: expected output for each row")

